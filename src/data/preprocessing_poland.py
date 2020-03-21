import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from contextlib import closing
from src.data.datasets import *
import numpy as np


def prepare_family_structure_from_voivodship_old(data_folder):
    voivodship = os.path.basename(data_folder)[0]
    families_per_household_df = pd.read_excel(os.path.join(data_folder, families_per_household_xlsx.file_name))

    family_structure_df = pd.read_excel(
        os.path.join(data_folder, os.pardir, voivodship, household_family_structure_old_xlsx.file_name),
        sheet_name=household_family_structure_old_xlsx.sheet_name)

    columns = family_structure_df.columns[1:]
    for col in columns:
        family_structure_df[col] = round(
            family_structure_df[col] * families_per_household_df.loc[0, col] / family_structure_df[col].sum())\
            .astype(int)

    family_structure_df.to_excel(os.path.join(data_folder, household_family_structure_old_xlsx.file_name), index=False)


def prepare_family_structure_from_voivodship(data_folder):
    voivodship = os.path.basename(data_folder)[0]
    voivodship_folder = os.path.join(data_folder, os.pardir, voivodship)
    df = pd.read_excel(os.path.join(voivodship_folder, household_family_structure_xlsx.file_name),
                       sheet_name=household_family_structure_xlsx.sheet_name)
    df2 = pd.melt(df,
                  id_vars=['family_type', 'relationship', 'house master', 'family_structure_regex'],
                  value_vars=[1, 2, 3, 4, 5, 6, 7], var_name='household_headcount',
                  value_name='probability_within_headcount')

    df2.to_excel(os.path.join(data_folder, household_family_structure_xlsx.file_name), index=False)


def generate_household_indices_old(data_folder):
    headcount2 = []
    family_type = []

    family_structure_df = pd.read_excel(os.path.join(data_folder, household_family_structure_old_xlsx.file_name))

    for i, row in family_structure_df.iterrows():
        headcount2.extend([row['Household size']] * int(row['One family']))
        family_type.extend([1] * row['One family'])
        headcount2.extend([row['Household size']] * int(row['Two families']))
        family_type.extend([2] * row['Two families'])
        headcount2.extend([row['Household size']] * int(row['Three families']))
        family_type.extend([3] * row['Three families'])
        headcount2.extend([row['Household size']] * int(row['One person']))
        family_type.extend([0] * row['One person'])
        headcount2.extend([row['Household size']] * int(row['Nonfamily']))
        family_type.extend([0] * row['Nonfamily'])

    household_df = pd.DataFrame(
        data={'headcount': headcount2, 'family_type': family_type, 'household_index': list(range(len(headcount2)))})
    household_df.set_index('household_index').to_excel(os.path.join(data_folder, households_old_xlsx.file_name))


def _generation_configuration_for_household(df, headcount, family_type, relationship, house_master):
    if family_type == 1:
        if house_master not in (np.nan, '', None):
            return df[(df.family_type == family_type) & (df.relationship == relationship)
                      & (df.house_master == house_master)]
        return df[(df.family_type == family_type) & (df.relationship == relationship)]
    if family_type in (2, 3):
        return df[(df.family_type == family_type)]
    if family_type == 0:
        if headcount == 1:
            return df[(df.family_type == family_type) & (df.relationship == 'Jednoosobowe')]
        return df[(df.family_type == family_type) & (df.relationship == 'Wieloosobowe')]
    raise ValueError(f'Unknown family type {family_type}')


def generate_household_indices(data_folder):
    household_headcount = []
    family_type = []
    relationship = []
    house_master = []
    family_structure_regex = []
    young = []
    middle = []
    elderly = []

    family_structure_df = pd.read_excel(os.path.join(data_folder, household_family_structure_xlsx.file_name),
                                        sheet_name=household_family_structure_xlsx.sheet_name)
    households_count_df = pd.read_excel(os.path.join(data_folder, households_count_xlsx.file_name),
                                        sheet_name=households_count_xlsx.sheet_name)
    generations_configuration_df = pd.read_excel(os.path.join(data_folder, generations_configuration_xlsx.file_name),
                                                 sheet_name=generations_configuration_xlsx.sheet_name)

    for i, hc_row in households_count_df.iterrows():
        # family structure given this headcount
        fs_df = family_structure_df[family_structure_df.household_headcount == hc_row.nb_of_people_in_household].copy()
        fs_df['total'] = (fs_df.probability_within_headcount * hc_row.nb_of_households).astype(int)

        for j, row in fs_df.iterrows():
            if row.total > 0:
                household_headcount.extend([row.household_headcount] * row.total)
                family_type.extend([row.family_type] * row.total)
                relationship.extend([row.relationship] * row.total)
                house_master.extend([row.house_master] * row.total)
                family_structure_regex.extend([row.family_structure_regex] * row.total)

                gc_df = _generation_configuration_for_household(generations_configuration_df, row.household_headcount,
                                                                row.family_type, row.relationship, row.house_master)
                gc_idx = np.random.choice(gc_df.index.tolist(), p=gc_df.probability, size=row.total)
                young.extend(gc_df.loc[gc_idx, 'young'])
                middle.extend(gc_df.loc[gc_idx, 'middle'])
                elderly.extend(gc_df.loc[gc_idx, 'elderly'])

    household_df = pd.DataFrame(data=dict(household_index=list(range(len(household_headcount))),
                                          household_headcount=household_headcount,
                                          family_type=family_type,
                                          relationship=relationship,
                                          house_master=house_master,
                                          family_structure_regex=family_structure_regex,
                                          young=young, middle=middle, elderly=elderly))

    household_df.set_index('household_index').to_excel(os.path.join(data_folder, households_xlsx.file_name))


def generate_generations_configuration(voivodship_folder, data_folder):
    v_config_df = pd.read_excel(os.path.join(voivodship_folder, generations_configuration_xlsx.file_name),
                                sheet_name='preprocessed', header=[0, 1])
    melted = pd.melt(v_config_df, id_vars=[('Unnamed: 0_level_0', 'family_type'),
                                           ('Unnamed: 1_level_0', 'relationship'),
                                           ('Unnamed: 2_level_0', 'house_master')],
                     var_name=['unit', 'category'],
                     value_name='total')
    melted = melted.rename(columns={('Unnamed: 0_level_0', 'family_type'): 'family_type',
                                    ('Unnamed: 1_level_0', 'relationship'): 'relationship',
                                    ('Unnamed: 2_level_0', 'house_master'): 'house_master'})
    melted['young'] = melted.category.isin(['cat1', 'cat4', 'cat5', 'cat7']).astype(int)
    melted['middle'] = melted.category.isin(['cat2', 'cat4', 'cat6', 'cat7']).astype(int)
    melted['elderly'] = melted.category.isin(['cat3', 'cat5', 'cat6', 'cat7']).astype(int)
    melted = melted[melted.category != 'total']
    melted = melted.drop(columns=['category'])
    melted['relationship'] = melted['relationship'].fillna('N/A')
    melted['house_master'] = melted['house_master'].fillna('N/A')
    pivoted = pd.pivot_table(melted, columns=['unit'], values='total',
                             index=['family_type', 'relationship', 'house_master', 'young', 'middle', 'elderly'],
                             aggfunc='first').reset_index()
    pivoted.households = pd.to_numeric(pivoted.households, errors='coerce')
    pivoted.people = pd.to_numeric(pivoted.people, errors='coerce')
    pivoted = pivoted.fillna(0)

    pivoted.loc[pivoted['family_type'] == 'Jednorodzinne', 'family_type'] = 1
    pivoted.loc[pivoted['family_type'] == 'Dwurodzinne', 'family_type'] = 2
    pivoted.loc[pivoted['family_type'] == 'Trzy i więcej rodzinne', 'family_type'] = 3
    pivoted.loc[pivoted['family_type'] == 'Nierodzinne', 'family_type'] = 0

    voivodship_workbook_path = os.path.join(voivodship_folder, generations_configuration_xlsx.file_name)
    book = load_workbook(voivodship_workbook_path)

    if generations_configuration_xlsx.sheet_name in book.sheetnames:
        del book[generations_configuration_xlsx.sheet_name]

    with closing(pd.ExcelWriter(voivodship_workbook_path, engine='openpyxl')) as writer:
        writer.book = book
        pivoted.to_excel(writer, sheet_name=generations_configuration_xlsx.sheet_name, index=False)
        writer.save()

    # update with probabilities
    df = pivoted.groupby(by=['family_type', 'relationship', 'house_master'])['households'].sum().reset_index()\
        .rename(columns={'households': 'total'})
    pivoted = pivoted.merge(df, how='left', on=['family_type', 'relationship', 'house_master'])
    pivoted['probability'] = pivoted['households'] / pivoted['total']

    output_file = os.path.join(data_folder, generations_configuration_xlsx.file_name)
    if os.path.isfile(output_file):
        os.unlink(output_file)

    pivoted.to_excel(output_file, sheet_name=generations_configuration_xlsx.sheet_name, index=False)


if __name__ == '__main__':
    project_dir = Path(__file__).resolve().parents[2]
    voivodship_folder = os.path.join(project_dir, 'data', 'processed', 'poland', 'D')
    data_folder = os.path.join(project_dir, 'data', 'processed', 'poland', 'DW')
    # prepare_family_structure_from_voivodship(data_folder)
    generate_household_indices(data_folder)
    # generate_generations_configuration(voivodship_folder, data_folder)
